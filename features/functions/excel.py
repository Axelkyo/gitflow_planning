import openpyxl

class funxl():

    def __init__(self, driver):
        self.driver = driver

    def getRowCount(file, path, sheetName):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return(sheet.max_row)

    def getColumnCount(file, path, sheetName):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return(sheet.max_column)

    def readData(file, path, sheetName, rowNum, colNum):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return sheet.cell(row = rowNum, column = colNum).value

    def writeData(file, path, sheetName, rowNum, colNum, data):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        sheet.cell(row = rowNum, column = colNum).value = data
        Workbook.save(path)




